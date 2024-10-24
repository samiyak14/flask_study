from flask import Flask, request, redirect, url_for, render_template, session
from flask_mail import Mail,Message
import openpyxl
import bcrypt


app = Flask(__name__)
app.secret_key = '0ec9e84f522344bda9ff60b99457800a'  

app.config['MAIL_SERVER']='smtp.gmail.com'
app.config['MAIL_PORT']=465
app.config['MAIL_USERNAME']='studentattendanceportal.cse@gmail.com'
app.config['MAIL_PASSWORD']='vdnd xmsy hcrl yeey'
app.config['MAIL_USE_TLS']=False
app.config['MAIL_USE_SSL']=True

mail=Mail(app)

def register_user(enrno, email, name, parent_email, password, role, subjects=''): 
    REGISTRATION_FILE = 'workbooks/registration_details.xlsx'
    wb = openpyxl.load_workbook(filename=REGISTRATION_FILE)
    if role == 'teacher':
        ws = wb['Teachers']
    else:
        ws = wb['Students']

    for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]== email or row[1]==email:
                return False

    hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
    if role == 'teacher':
        ws.append([email, name, hashed_password.decode(), role, subjects])
    else:
        ws.append([enrno, email, name, parent_email, hashed_password.decode(), role])
    wb.save(filename=REGISTRATION_FILE)
    return True

def login_user(email, password, role):
    REGISTRATION_FILE = 'workbooks/registration_details.xlsx'
    wb = openpyxl.load_workbook(filename=REGISTRATION_FILE)
    if role == 'teacher':
        ws = wb['Teachers']
    else:
        ws = wb['Students']
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == email:
            stored_hashed_password = row[2].encode()
            if bcrypt.checkpw(password.encode(), stored_hashed_password):
                return row  # Return the whole row for further use
        if row[1] == email:
            stored_hashed_password = row[4].encode()
            if bcrypt.checkpw(password.encode(), stored_hashed_password):
                return row  
    return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login_student', methods=['GET', 'POST'])
def login_student():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user_data = login_user(email, password, 'student')
        if user_data:
            session['email'] = email
            session['role'] = 'student'
            session['name'] = user_data[2]  
            return redirect(url_for('student_dashboard'))
        return 'Invalid credentials.'
    return render_template('login_student.html')

@app.route('/login_teacher', methods=['GET', 'POST'])
def login_teacher():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user_data = login_user(email, password, 'teacher')
        if user_data:
            session['email'] = email
            session['role'] = 'teacher'
            session['name'] = user_data[1]  # Extract name
            session['subjects'] = user_data[4]  # Extract subjects
            return redirect(url_for('teacher_dashboard'))
        return 'Invalid credentials'
    return render_template('login_teacher.html')

@app.route('/register_student', methods=['GET', 'POST'])
def register_student():
    if request.method == 'POST':
        enrno = request.form['enrno']
        email = request.form['email']
        name = request.form['name']
        parent_email = request.form['parent_email']
        password = request.form['password']
        if register_user(enrno, email, name, parent_email, password, 'student') is True:
            return redirect(url_for('login_student'))
        else:
            return 'email exists.'

    return render_template('register_student.html')

@app.route('/register_teacher', methods=['GET', 'POST'])
def register_teacher():
    if request.method == 'POST':
        email = request.form['email']
        name = request.form['name']
        password = request.form['password']
        subjects = request.form['subjects']
        if register_user('', email, name, '', password, 'teacher', subjects) is True:
            return redirect(url_for('login_teacher'))
        else :
            return 'Email already exists.'
    return render_template('register_teacher.html')

@app.route('/teacher_dashboard')
def teacher_dashboard():
    if 'role' in session and session['role'] == 'teacher':
        subjects = session.get('subjects', '').split(', ')
        return render_template('teacher_dashboard.html', subjects=subjects)
    return redirect(url_for('index'))

@app.route('/select_subject_view_form')
def select_subject_view_form():
    if 'role' in session and session['role']=='teacher':
        subjects=session.get('subjects','').split(', ')
        return render_template('select_subject_view.html', subjects=subjects)
    return redirect(url_for('index'))

@app.route('/select_subject_view', methods=['POST'])
def select_subject_view():
    if 'role' in session and session['role'] == 'teacher':
        selected_class = request.form['class']
        selected_subject = request.form['subject']
        return redirect(url_for('view_attendance', selected_class=selected_class, selected_subject=selected_subject))
    return redirect(url_for('index'))

@app.route('/view_attendance/<selected_class>/<selected_subject>', methods=['GET', 'POST'])
def view_attendance(selected_class, selected_subject):
    if 'role' in session and session['role'] == 'teacher':
        workbook_name = f'workbooks/{selected_class}.xlsx'
        
        try:
            book = openpyxl.load_workbook(workbook_name, data_only=True)
            ws = book[selected_subject]
            sheet = book['Attendance']

            attendance_data = []  #this is going to be a list of tuples.

            subject_column_index = None
            for col in range(5, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == selected_subject:
                    subject_column_index = col
                    break

            # Collect attendance data
            if subject_column_index is not None:
                for row in range(2, sheet.max_row + 1):
                    roll_no = sheet.cell(row=row, column=1).value
                    student_name = sheet.cell(row=row, column=4).value
                    attendance_percentage = sheet.cell(row=row, column=subject_column_index).value
                    if student_name and attendance_percentage is not None:
                        attendance_data.append((roll_no, student_name, attendance_percentage))
            print(attendance_data)

            # Count absences by day of the week for average calculation
            total_absentee_count = {
                'Monday': 0,
                'Tuesday': 0,
                'Wednesday': 0,
                'Thursday': 0,
                'Friday': 0,
                'Saturday': 0
            }

            total_students_per_day = {
                'Monday': 0,
                'Tuesday': 0,
                'Wednesday': 0,
                'Thursday': 0,
                'Friday': 0,
                'Saturday': 0
            }

            # Loop through the days and count absentees
            for col in range(5, ws.max_column + 1):
                day_of_week = ws.cell(row=3, column=col).value
                if day_of_week in total_absentee_count:
                    for row in range(4, ws.max_row + 1):
                        attendance_status = ws.cell(row=row, column=col).value
                        if attendance_status == 'A':  
                            total_absentee_count[day_of_week] += 1
                        
                        total_students_per_day[day_of_week] += 1
            
            # Calculate average percentages
            average_percentage_absentees = {
                day: round((total_absentee_count[day] / total_students_per_day[day] * 100), 3) if total_students_per_day[day] > 0 else 0
                for day in total_absentee_count
            }

            # Convert average_percentage_absentees to lists for Chart.js
            labels = list(average_percentage_absentees.keys())
            data = list(average_percentage_absentees.values())

            print("Labels:", labels)
            print("Data:", data)

            return render_template('view_attendance.html', 
                                   selected_class=selected_class, 
                                   selected_subject=selected_subject, 
                                   attendance_data=attendance_data,
                                   labels=labels,
                                   data=data)
        except Exception as e:
            return render_template('error.html', message=f'Error occurred: {e}')

    return redirect(url_for('index'))


@app.route('/parent_emails', methods=['GET', 'POST'])
def parent_emails():
    wbr = openpyxl.load_workbook(filename='workbooks/registration_details.xlsx')
    sheet_s = wbr['Students']
    s_details = []  # store all the enrollment numbers of registered students
    for row in sheet_s.iter_rows(min_row=2, values_only=True):
        s_details.append(row[0])
    print(s_details)

    defaulter_list1 = []  # stores the registered students' enrollment numbers with less attendance
    defaulter_list = []  
    
    for enr in s_details:
        if check_enrollment_exists_SE(enr):
            wbSE = openpyxl.load_workbook(filename='workbooks/SE.xlsx', data_only=True)
            sheet1 = wbSE['Attendance']

            for row in sheet1.iter_rows(min_row=2, values_only=True):
                attendance_value = row[14]
                if enr == row[1] and isinstance(attendance_value, (int, float)) and attendance_value < 75.00:
                    defaulter_list.append(enr)
                    defaulter_list1.append((enr, row[3], attendance_value))

        elif check_enrollment_exists_TE(enr):
            wbSE = openpyxl.load_workbook(filename='workbooks/TE.xlsx', data_only=True)
            sheet1 = wbSE['Attendance']

            for row in sheet1.iter_rows(min_row=2, values_only=True):
                attendance_value = row[15]
                if enr == row[1] and isinstance(attendance_value, (int, float)) and attendance_value < 75.00:
                    defaulter_list.append(enr)
                    defaulter_list1.append((enr, row[3], attendance_value))

    print(defaulter_list)
    d_details = []  # stores the name and parent's email of students with less attendance for sending formatted emails.

    for enrno in defaulter_list:
        for row in sheet_s.iter_rows(min_row=2, values_only=True):
            if enrno == row[0]:
                d_details.append({'name': row[2], 'parents_mail_id': row[3]})

    if request.method == 'POST':
        try:
            for student in d_details:
                name = student['name']
                recipient = student['parents_mail_id']
                email_template = """Dear Parent,
We hope this email finds you well. We are writing to inform you that your child, {name}, has an average attendance rate below the mandatory 75%.
Regular attendance is crucial for his/her academic success, and we encourage you to discuss the importance of attending classes regularly. If the attendance continues to be low, strict measures may need to be taken as per the university norms.
Please feel free to reach out if you have any questions or require further information.
Best regards,
Department of CSE (AI-ML)
Finolex Academy of Management and Technology
"""

                personalised_email = email_template.format(name=name)
                msg = Message(f"{name}'s Attendance below 75%", sender='studentattendanceportal.cse@gmail.com', recipients=[recipient])
                msg.body = personalised_email

                mail.send(msg)

            return render_template('success.html', message="Emails sent successfully.")

        except Exception as e:
            return f"Failed to send email: {str(e)}"

    return render_template('parent_emails.html', defaulter_list1=defaulter_list1)


def get_student_total_attendance_TE(enrno):
    wb = openpyxl.load_workbook(filename='workbooks/TE.xlsx',data_only=True)
    ws = wb['Attendance']
    

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == enrno: 
            att=[row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15]]
            return att 

    return None  # Return None if the student is not found

def get_student_total_attendance_SE(enrno):
    wb = openpyxl.load_workbook(filename='workbooks/SE.xlsx',data_only=True)
    ws = wb['Attendance']  
    

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == enrno:  
            att=[row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14]]
            return att  
    return None  # Return None if the student is not found

def check_enrollment_exists_SE(enrollment_number, filename='SE.xlsx'):

    wb = openpyxl.load_workbook(filename='workbooks/SE.xlsx',data_only=True)
    ws = wb['Attendance'] 

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == enrollment_number: 
            return True  # Enrollment number found
    return False  # Enrollment number not found

def check_enrollment_exists_TE(enrollment_number, filename='TE.xlsx'):

    wb = openpyxl.load_workbook(filename='workbooks/TE.xlsx',data_only=True)
    ws = wb['Attendance'] 

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == enrollment_number: 
            return True  

    return False  


@app.route('/student_dashboard')
def student_dashboard():
    wb = openpyxl.load_workbook(filename='workbooks/registration_details.xlsx')
    ws = wb['Students']
    email = session.get('email')
    enrno = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == email:  
            enrno = row[0]  
            break  # Exit loop once found
    
    if enrno is None:
        return 'Student not found.'  # Handle case where email does not match any student

    if 'role' in session and session['role'] == 'student':
        if check_enrollment_exists_SE(enrno):
            att = get_student_total_attendance_SE(enrno)
            if att is not None:
                return render_template('student_dashboard_SE.html', 
                                       EMIII="{:.2f}".format(att[0]),
                                       DS="{:.2f}".format(att[1]),
                                       DLCOA="{:.2f}".format(att[2]),
                                       CG="{:.2f}".format(att[3]),
                                       OOPM="{:.2f}".format(att[4]),
                                       DSGT="{:.2f}".format(att[5]),
                                       OOPMLAB="{:.2f}".format(att[6]),
                                       DSLAB="{:.2f}".format(att[7]),
                                       CGLAB="{:.2f}".format(att[8]),
                                       DLCOALAB="{:.2f}".format(att[9]),
                                       total_attendance="{:.2f}".format(att[10]),
                                       class_name='SE')
            else:
                return 'Attendance not found for the student.'
        elif check_enrollment_exists_TE(enrno):
            att = get_student_total_attendance_TE(enrno)
            if att is not None:
                return render_template('student_dashboard_TE.html', 
                                       CN="{:.2f}".format(att[0]),
                                       WC="{:.2f}".format(att[1]),
                                       AI="{:.2f}".format(att[2]),
                                       DWHM="{:.2f}".format(att[3]),
                                       DLOC="{:.2f}".format(att[4]),
                                       IOT="{:.2f}".format(att[5]),
                                       BCE="{:.2f}".format(att[6]),
                                       WCLAB="{:.2f}".format(att[7]),
                                       AILAB="{:.2f}".format(att[8]),
                                       DWHMLAB="{:.2f}".format(att[9]),
                                       BCELAB="{:.2f}".format(att[10]),
                                       total_attendance="{:.2f}".format(att[11]),
                                       class_name='TE')
            else:
                return 'Attendance not found for the student.'

    return redirect(url_for('index'))
@app.route('/select_subject_form')
def select_subject_form():
    if 'role' in session and session['role'] == 'teacher':
        subjects = session.get('subjects', '').split(', ')  # Retrieve subjects from session
        return render_template('select_subject.html', subjects=subjects) 
    return redirect(url_for('index'))

@app.route('/select_subject', methods=['POST'])
def select_subject():
    if 'role' in session and session['role'] == 'teacher':
        selected_class = request.form['class']
        selected_subject = request.form['subject']
        return redirect(url_for('attendance_form', selected_class=selected_class, selected_subject=selected_subject))
    return redirect(url_for('index'))

@app.route('/attendance_form/<selected_class>/<selected_subject>', methods=['GET', 'POST'])
def attendance_form(selected_class, selected_subject):
    if 'role' in session and session['role'] == 'teacher':
        if request.method == 'POST':
            marking_method = request.form['marking_method']  
            roll_numbers_input = request.form['roll_numbers'].split(',')
            roll_numbers_input = set(int(roll.strip()) for roll in roll_numbers_input)
            date = request.form['date']
            day = request.form['day']

            workbook_name = f'workbooks/{selected_class}.xlsx'
            division = selected_subject

            try:
                book = openpyxl.load_workbook(workbook_name)
                ws = book[division]

                current_column = 1
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value is None:
                            break

                current_column = cell.column
                roll_number_column = 1

                
                for row in range(4, ws.max_row + 1):
                    roll_number = ws.cell(row=row, column=roll_number_column).value
                    cell = ws.cell(row=row, column=current_column)
                    
                    if marking_method == 'absent':
                        # If marking by absent roll numbers
                        if roll_number in roll_numbers_input:
                            cell.value = 'A'  
                        else:
                            if cell.value is None:
                                cell.value = 'P'  
                    else:
                        # If marking by present roll numbers
                        if roll_number in roll_numbers_input:
                            cell.value = 'P' 
                        else:
                            if cell.value is None:
                                cell.value = 'A'  

                # Add the date and day in the top row of the column
                ws.cell(row=2, column=current_column).value = date
                ws.cell(row=3, column=current_column).value = day

                book.save(workbook_name)

                return render_template('success.html', message='Attendance updated successfully!')
            except Exception as e:
                return render_template('error.html', message=f'Error occurred: {e}')
        
        return render_template('attendance_form.html', selected_class=selected_class, selected_subject=selected_subject)
    return redirect(url_for('index'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
